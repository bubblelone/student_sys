from django.shortcuts import render, redirect
#from .models import Student
#from .forms import StudentForm , QueryForm, UploadFileForm
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import load_workbook




def excleToDb():
    excel = load_workbook('../1.xlsx')
    table = excel.get_sheet_by_name('Sheet1')

    for row in table.iter_rows(2,3,1,6):
        dic = ['name', 'sex', 'profession', 'email', 'qq', 'phone']
        tmplist = []
        for cell in row:
            #if cell == '1' or cell == '2':
            #    cell = int(cell)
            tmplist.append(cell)
        tmpdic =dict(zip(dic, tmplist))
        print(tmpdic)
        #Student.objects.create(**tmpdic)

excleToDb()