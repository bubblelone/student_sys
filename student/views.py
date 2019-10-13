from django.shortcuts import render, redirect
from .models import Student
from .forms import StudentForm , QueryForm, UploadFileForm
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.template.context_processors import csrf
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate

# Create your views here.

@login_required
def index(request):
    username = request.user
    students = Student.objects.all().order_by('-created_time')
    formQuery = QueryForm()
    uploadform = UploadFileForm()


    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            cleaned_data = form.cleaned_data
            student = Student()
            student.name = cleaned_data['name']
            student.sex = cleaned_data['sex']
            student.email = cleaned_data['email']
            student.profession = cleaned_data['profession']
            student.qq = cleaned_data['qq']
            student.phone = cleaned_data['phone']
            student.save()
            return redirect(reverse('index'))
    else:
        form = StudentForm()

    context = {
            'students': students,
            'form': form,
        'formQuery': formQuery,
        'uploadform': uploadform,
        'username': username,
        }
    return render(request, 'index.html', context=context)

def search_name(request):


    formQuery = QueryForm(request.POST)

    form = StudentForm()

    if formQuery.is_valid():
        cleaned_data = formQuery.cleaned_data
        search_name = cleaned_data['name']
        search_profession = cleaned_data['profession']
        search_email = cleaned_data['email']
        search_sex = cleaned_data['sex']
        search_qq = cleaned_data['qq']
        search_phone = cleaned_data['phone']


        if search_sex == None or search_sex == '':
            students = Student.objects.filter(name__contains=search_name,
                                              profession__contains=search_profession,

                                              qq__contains=search_qq,
                                              phone__contains=search_phone,
                                              email__contains=search_email,
                                              ).order_by('-created_time')
        else:
            students = Student.objects.filter(name__contains=search_name, sex=search_sex,
                                              profession__contains=search_profession,

                                              qq__contains=search_qq,
                                              phone__contains=search_phone,
                                              email__contains=search_email,
                                              ).order_by('-created_time')

    context = {
            'students': students,
            'form': form,
            'formQuery': formQuery
        }
    return render(request, 'index.html', context=context)

def hello(request):
    return render(request, 'hello.html')

def hello2(request):
    return render(request, 'hello2.html')

def hello3(request):
    return render(request, 'hello3.html')

def hello4(request):
    return render(request, 'hello4.html')

def login3(request):
    return render(request, 'login3.html')

def list1(request):
    return HttpResponse('这是列表1')

def list2(request):
    return HttpResponse('这是列表2')

def excel_upload(request):

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        f = request.FILES['file']

        type_excel = f.name.split('.')[1]
        if form.is_valid() and 'xlsx' == type_excel:
            #handle_uploaded_file(request.FILES['file'])
            excleToDb(f)
            return render(request, 'hello.html')

#@csrf_exempt
def excel_upload2(request):

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        f = request.FILES['file']

        type_excel = f.name.split('.')[1]
        if form.is_valid() and 'jpg' == type_excel:
            handle_uploaded_file(request.FILES['file'])
            #excleToDb(f)
            return JsonResponse({"code":0, "msg":"success", "data":{"src":"http://123.jpg"}})
            #return 'ok'
            #return render(request, 'hello.html')


def handle_uploaded_file(f):
    with open('3.jpg', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

def excleToDb(f):
    excel = load_workbook(f)
    table = excel.get_sheet_by_name('Sheet1')
    for row in table.iter_rows(2,100,1,6):
        dic = ['name', 'sex', 'profession', 'email', 'qq', 'phone']
        tmplist = []
        for cell in row:

            tmplist.append(cell.value)
        tmpdic =dict(zip(dic, tmplist))
        print(tmpdic)
        Student.objects.create(**tmpdic)


def get_csrf(request):
    x = csrf(request)
    csrf_token = x['csrf_token']
    return HttpResponse(csrf_token)









